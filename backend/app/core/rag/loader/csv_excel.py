from __future__ import annotations

import csv
import io
from app.core.common.logging import TerraLogUtil as logger
from pathlib import Path
from typing import ClassVar

from app.core.rag.loader.base import BaseLoader
from app.core.rag.models.document import Document, Metadata



class CSVLoader(BaseLoader):
    supported_extensions: ClassVar[list[str]] = [".csv"]

    def __init__(self, encoding: str = "utf-8"):
        self._encoding = encoding

    def load(self, source: str | Path | bytes, **kwargs) -> list[Document]:
        raw = self._read_bytes(source)
        text = raw.decode(self._encoding, errors="replace")
        source_name = self._resolve_source_name(source)

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []

        header = rows[0]
        documents: list[Document] = []
        for i, row in enumerate(rows[1:], start=1):
            pairs = [f"{h}: {v}" for h, v in zip(header, row) if v.strip()]
            if pairs:
                documents.append(
                    Document(
                        content="\n".join(pairs),
                        metadata=Metadata(
                            source=source_name,
                            file_type="csv",
                            extra={"row_number": i},
                        ),
                    )
                )
        logger.info(
            "Loaded CSV: source=%s, docs=%d",
            source_name, len(documents),
        )
        return documents


class ExcelLoader(BaseLoader):
    """Excel loader using openpyxl."""

    supported_extensions: ClassVar[list[str]] = [".xlsx", ".xls"]

    def load(self, source: str | Path | bytes, **kwargs) -> list[Document]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel loading. Install with: pip install fastrag[excel]")

        raw = self._read_bytes(source)
        source_name = self._resolve_source_name(source)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

        documents: list[Document] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(c) if c is not None else "" for c in rows[0]]
            for i, row in enumerate(rows[1:], start=1):
                values = [str(c) if c is not None else "" for c in row]
                pairs = [f"{h}: {v}" for h, v in zip(header, values) if v.strip()]
                if pairs:
                    documents.append(
                        Document(
                            content="\n".join(pairs),
                            metadata=Metadata(
                                source=source_name,
                                file_type="excel",
                                extra={"sheet": sheet_name, "row_number": i},
                            ),
                        )
                    )

        sheet_count = len(wb.sheetnames)
        wb.close()
        logger.info(
            "Loaded Excel: source=%s, sheets=%d, docs=%d",
            source_name, sheet_count, len(documents),
        )
        return documents
